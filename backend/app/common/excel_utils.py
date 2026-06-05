from typing import Any, Dict, List, Optional
import openpyxl


def read_excel(
    file_path: str, sheet_name: Optional[str] = None
) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        rows.append(row_dict)
    wb.close()
    return rows


def write_excel(
    data: List[Dict[str, Any]],
    file_path: str,
    headers: Optional[List[str]] = None,
):
    if not data:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    if headers is None:
        headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
    wb.save(file_path)


def validate_columns(
    headers: List[str], required_columns: List[str]
) -> List[str]:
    return [col for col in required_columns if col not in headers]
